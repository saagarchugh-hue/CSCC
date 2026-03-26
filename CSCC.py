from __future__ import annotations

import csv
from openpyxl import load_workbook

INPUT_FILE = "CA managed merchants.xlsx"
OUTPUT_FILE = "merchant_success_command_center.csv"
DEFAULT_DAY = "01"

MONTH_LABELS = {
    "2026-04": "Apr 2026",
    "2026-05": "May 2026",
    "2026-06": "Jun 2026",
    "2026-07": "Jul 2026",
    "2026-08": "Aug 2026",
    "2026-09": "Sep 2026",
    "2026-10": "Oct 2026",
    "2026-11": "Nov 2026",
    "2026-12": "Dec 2026",
    "2027-01": "Jan 2027",
    "2027-02": "Feb 2027",
    "2027-03": "Mar 2027",
}

PEAK_MAP = {
    "spring_home": ["Apr", "May"],
    "summer_home": ["Jun", "Jul"],
    "bts": ["Aug", "Sep"],
    "labour_day": ["Sep"],
    "holiday": ["Nov", "Dec"],
    "boxing_day": ["Dec"],
    "new_year": ["Jan", "Feb"],
    "pre_summer_aesthetic": ["Apr", "May", "Jun"],
    "auto_spring": ["Apr", "May", "Jun"],
    "auto_winter_prep": ["Oct", "Nov"],
}

OVERRIDES = {
    "Samsung": {"vertical": "Electronics & Wireless", "tier": "Tier 0", "seasonality": ["bts", "holiday"]},
    "Cozey": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "Fizz": {"vertical": "Electronics & Wireless", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
    "Structube": {"vertical": "Furniture & Home", "tier": "Tier 0", "seasonality": ["spring_home", "bts", "holiday"]},
    "Rona": {"vertical": "Home Improvement", "tier": "Tier 0", "seasonality": ["spring_home", "summer_home", "holiday"]},
    "Nuovo Photography": {"vertical": "Photography & Services", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Endy Sleep": {"vertical": "Sleep", "tier": "Tier 1", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Footlocker": {"vertical": "Retail / Multi", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
    "Thermomix": {"vertical": "Appliances", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Goodmorning.com": {"vertical": "Sleep", "tier": "Tier 0", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Silk and Snow": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Skin Vitality": {"vertical": "Aesthetics / Medical", "tier": "Tier 1", "seasonality": ["new_year", "pre_summer_aesthetic", "holiday"]},
    "Canadian Beauty School": {"vertical": "Education", "tier": "Tier 2", "seasonality": ["bts", "new_year"]},
    "Jump+": {"vertical": "Electronics & Wireless", "tier": "Tier 1", "seasonality": ["bts", "holiday"]},
    "LG Canada": {"vertical": "Appliances", "tier": "Tier 1", "seasonality": ["spring_home", "holiday"]},
    "Removery": {"vertical": "Aesthetics / Medical", "tier": "Tier 2", "seasonality": ["new_year", "pre_summer_aesthetic"]},
    "Sleep Country": {"vertical": "Sleep", "tier": "Tier 0", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Clutch": {"vertical": "Auto", "tier": "Tier 1", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Oliver Jewellery": {"vertical": "Jewelry", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Cellcom Communications": {"vertical": "Electronics & Wireless", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
    "Canac": {"vertical": "Home Improvement", "tier": "Tier 1", "seasonality": ["spring_home", "summer_home", "holiday"]},
    "PMC Tire": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Canada Wheels": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Northern Fitness": {"vertical": "Fitness & Mobility", "tier": "Tier 2", "seasonality": ["new_year", "holiday"]},
    "Parts Engine": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Urban Barn Ltd.": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "Browns Shoes": {"vertical": "Retail / Multi", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
    "Urban Barn": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "Linen Chest": {"vertical": "Furniture & Home", "tier": "Tier 0", "seasonality": ["spring_home", "holiday", "boxing_day"]},
    "Tdot Performance": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "MedSpa Partners": {"vertical": "Aesthetics / Medical", "tier": "Tier 1", "seasonality": ["new_year", "pre_summer_aesthetic", "holiday"]},
    "EQ3": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "Canon": {"vertical": "Electronics & Wireless", "tier": "Tier 1", "seasonality": ["bts", "holiday"]},
    "Camera Canada": {"vertical": "Photography & Services", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Sparx Hockey": {"vertical": "Sports & Outdoor", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Excellent Photo": {"vertical": "Photography & Services", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Mobile Klinik": {"vertical": "Electronics & Wireless", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
    "Bath Depot In-Store": {"vertical": "Home Improvement", "tier": "Tier 2", "seasonality": ["spring_home", "summer_home"]},
    "Bath Depot": {"vertical": "Home Improvement", "tier": "Tier 2", "seasonality": ["spring_home", "summer_home"]},
    "Lee Valley Tools Limited": {"vertical": "Home Improvement", "tier": "Tier 2", "seasonality": ["spring_home", "holiday"]},
    "Sunday Furniture Ltd": {"vertical": "Furniture & Home", "tier": "Tier 2", "seasonality": ["spring_home", "bts", "holiday"]},
    "Hush Blankets": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["holiday"]},
    "GE Applicances": {"vertical": "Appliances", "tier": "Tier 1", "seasonality": ["spring_home", "holiday"]},
    "Rove Concepts": {"vertical": "Furniture & Home", "tier": "Tier 2", "seasonality": ["spring_home", "bts", "holiday"]},
    "JYSK": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "The Camera Store": {"vertical": "Photography & Services", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Giant Bicycle Canada": {"vertical": "Fitness & Mobility", "tier": "Tier 2", "seasonality": ["spring_home", "summer_home"]},
    "Ashley Home Store": {"vertical": "Furniture & Home", "tier": "Tier 2", "seasonality": ["spring_home", "bts", "holiday"]},
    "Casper": {"vertical": "Sleep", "tier": "Tier 1", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Nissan Stores": {"vertical": "Auto", "tier": "Tier 1", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "iDrinkCoffee": {"vertical": "Appliances", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Bells of Steel": {"vertical": "Fitness & Mobility", "tier": "Tier 2", "seasonality": ["new_year", "holiday"]},
    "Polysleep": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "West Coast Kids": {"vertical": "Retail / Multi", "tier": "Tier 2", "seasonality": ["holiday"]},
    "Emma Sleep": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "GhostBed": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "Tire Connect": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Wheelwiz": {"vertical": "Auto", "tier": "Tier 2", "seasonality": ["auto_spring", "auto_winter_prep"]},
    "Transformer Table": {"vertical": "Furniture & Home", "tier": "Tier 2", "seasonality": ["spring_home", "bts", "holiday"]},
    "Trevi": {"vertical": "Outdoor & Pools", "tier": "Tier 2", "seasonality": ["summer_home"]},
    "Maison Corbeil": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "MUST": {"vertical": "Furniture & Home", "tier": "Tier 1", "seasonality": ["spring_home", "bts", "holiday"]},
    "Jardin De Ville": {"vertical": "Outdoor & Pools", "tier": "Tier 2", "seasonality": ["summer_home"]},
    "Hudson's Bay": {"vertical": "Retail / Multi", "tier": "Tier 1", "seasonality": ["bts", "holiday"]},
    "Royal Distributing": {"vertical": "Sports & Outdoor", "tier": "Tier 2", "seasonality": ["spring_home", "holiday"]},
    "Simba Sleep": {"vertical": "Sleep", "tier": "Tier 2", "seasonality": ["spring_home", "labour_day", "holiday"]},
    "The Sony Shop": {"vertical": "Electronics & Wireless", "tier": "Tier 2", "seasonality": ["bts", "holiday"]},
}

def normalize_name(name: str) -> str:
    return " ".join(str(name).strip().split())

def infer_vertical_and_seasonality(name: str):
    if name in OVERRIDES:
        item = OVERRIDES[name]
        return item["vertical"], item["tier"], item["seasonality"]

    n = name.lower()

    if any(k in n for k in ["sleep", "mattress", "ghostbed", "polysleep", "novosbed", "simba"]):
        return "Sleep", "Tier 2", ["spring_home", "labour_day", "holiday"]

    if any(k in n for k in ["nissan", "tire", "motorsport", "diesel", "lube", "speedy glass", "vitroplus", "wheel", "parts ", "car and truck", "subie", "performance", "auto"]):
        return "Auto", "Tier 2", ["auto_spring", "auto_winter_prep"]

    if any(k in n for k in ["spa", "aesthetic", "cosmetic", "removery", "urospot", "skin vitality", "new look", "opht", "optique", "vision", "optical", "eyecare", "iris", "vogue optical", "hearinglife"]):
        return "Aesthetics / Medical", "Tier 2", ["new_year", "pre_summer_aesthetic", "holiday"]

    if any(k in n for k in ["furniture", "barn", "chest", "corbeil", "sofa", "must", "dufresne", "jysk", "table"]):
        return "Furniture & Home", "Tier 2", ["spring_home", "bts", "holiday"]

    if any(k in n for k in ["bath depot", "rona", "canac", "lee valley", "tenaquip"]):
        return "Home Improvement", "Tier 2", ["spring_home", "summer_home", "holiday"]

    if any(k in n for k in ["camera", "canon", "sony", "sigma", "drone", "photo"]):
        return "Photography & Services", "Tier 2", ["holiday"]

    if any(k in n for k in ["samsung", "mobile", "wireless", "jump+", "rebelo", "gaming pc", "electronics", "bowers", "denon", "marantz", "polk", "definitive technology", "rewireless", "rwireless"]):
        return "Electronics & Wireless", "Tier 2", ["bts", "holiday"]

    if any(k in n for k in ["fitness", "weights", "bicycle", "bike", "hockey", "bells of steel", "rad power"]):
        return "Fitness & Mobility", "Tier 2", ["new_year", "spring_home", "holiday"]

    if any(k in n for k in ["jeweller", "diamonds", "bijoux", "luxedujour", "matt and nat"]):
        return "Jewelry", "Tier 2", ["holiday"]

    if any(k in n for k in ["shoes", "softmoc", "footlocker", "bay", "giant tiger", "hart", "indochino", "ren's pet", "showpass"]):
        return "Retail / Multi", "Tier 2", ["bts", "holiday"]

    if any(k in n for k in ["coffee", "kitchen", "thermomix", "roborock", "narwal", "appliances", "lg", "ge"]):
        return "Appliances", "Tier 2", ["spring_home", "holiday"]

    if any(k in n for k in ["school", "academy", "wizeprep", "conduite", "programme", "block"]):
        return "Education", "Tier 2", ["bts", "new_year"]

    return "Other", "Tier 2", ["holiday"]

def seasonality_to_peak_months(seasonality_codes):
    months = []
    for code in seasonality_codes:
        months.extend(PEAK_MAP.get(code, []))
    deduped = []
    for m in months:
        if m not in deduped:
            deduped.append(m)
    return deduped

def month_to_label(month_key: str) -> str:
    return MONTH_LABELS[month_key]

def peak_month_to_pre_engagement_months(peak_mon: str):
    mapping = {
        "Jan": ["2026-10", "2026-11"],
        "Feb": ["2026-11", "2026-12"],
        "Mar": ["2026-12", "2027-01"],
        "Apr": ["2027-01", "2027-02"],
        "May": ["2027-02", "2027-03"],
        "Jun": ["2026-04"],
        "Jul": ["2026-04", "2026-05"],
        "Aug": ["2026-05", "2026-06"],
        "Sep": ["2026-06", "2026-07"],
        "Oct": ["2026-07", "2026-08"],
        "Nov": ["2026-08", "2026-09"],
        "Dec": ["2026-09", "2026-10"],
    }
    return mapping.get(peak_mon, [])

def engagement_type_for_month(month_key: str, peak_mon: str):
    pre = peak_month_to_pre_engagement_months(peak_mon)
    if len(pre) == 2:
        if month_key == pre[0]:
            return "T-12 to T-8 planning"
        if month_key == pre[1]:
            return "T-8 to T-4 launch prep"
    if len(pre) == 1 and month_key == pre[0]:
        return "Peak prep"
    return "Seasonal engagement"

def playbook_for_vertical(vertical: str):
    mapping = {
        "Furniture & Home": "Spring refresh / move season / holiday big-ticket financing",
        "Sleep": "Spring / Labour Day / holiday mattress and bedding push",
        "Appliances": "Spring reno / holiday appliance promo planning",
        "Electronics & Wireless": "BTS / holiday device and accessory financing",
        "Fitness & Mobility": "New Year / spring / holiday equipment financing",
        "Auto": "Spring upgrade / winter prep parts and service financing",
        "Outdoor & Pools": "Pre-summer install and outdoor living push",
        "Retail / Multi": "BTS / holiday merchandising and offer coordination",
        "Aesthetics / Medical": "New Year / pre-summer treatment package financing",
        "Home Improvement": "Spring reno / summer outdoor / holiday tools",
        "Photography & Services": "Holiday gifting and creator equipment spikes",
        "Jewelry": "Holiday gifting and event-driven conversion",
        "Education": "BTS / New Year enrollment financing",
        "Other": "General seasonal planning cadence",
    }
    return mapping.get(vertical, "General seasonal planning cadence")

def leadership_flag(tier: str, priority: str):
    return "Yes" if tier == "Tier 0" or priority == "Critical" else "No"

def next_action_for_priority(priority: str):
    if priority == "Critical":
        return "Confirm promo economics and homepage placement"
    if priority == "High":
        return "Lock creative and launch checklist"
    if priority == "Medium":
        return "Schedule planning touchpoint"
    return "Monitor"

# Managed merchants sheet layout: 0=Account (merchant), 1=CSM, 2=FY26 FC GMV
CSM_COLUMN_INDEX = 1
GMV_COLUMN_INDEX = 2


def format_fy26_gmv(value) -> str:
    """Format FY26 FC GMV for display (currency or pass-through string)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        v = float(value)
        if abs(v) >= 1_000_000:
            return f"${v / 1_000_000:.2f}M"
        if abs(v) >= 1_000:
            return f"${v:,.0f}"
        return f"${v:,.2f}"
    s = str(value).strip()
    return s


def read_merchants_from_xlsx(path: str) -> tuple[list[str], dict[str, str], dict[str, str]]:
    """
    Returns (merchant_names, merchant_to_owner_map, merchant_to_fy26_gmv).
    CSM from column 1; FY26 FC GMV from column 2 when present.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    names = []
    seen = set()
    merchant_to_owner: dict[str, str] = {}
    merchant_to_gmv: dict[str, str] = {}

    for row in ws.iter_rows(values_only=True):
        value = row[0]
        if value is None:
            continue
        name = normalize_name(value)
        if not name or name.lower() == "all merchants" or name.lower() == "new merchants":
            continue
        if name not in seen:
            seen.add(name)
            names.append(name)
            csm = row[CSM_COLUMN_INDEX] if len(row) > CSM_COLUMN_INDEX else None
            merchant_to_owner[name] = normalize_name(str(csm)) if csm is not None and str(csm).strip() else ""
            gmv_raw = row[GMV_COLUMN_INDEX] if len(row) > GMV_COLUMN_INDEX else None
            merchant_to_gmv[name] = format_fy26_gmv(gmv_raw) if gmv_raw is not None and str(gmv_raw).strip() != "" else ""
    return names, merchant_to_owner, merchant_to_gmv


def build_rows(
    merchant_names: list[str],
    merchant_to_owner: dict[str, str],
    merchant_to_gmv: dict[str, str] | None = None,
):
    gmv_map = merchant_to_gmv or {}
    rows = []
    for merchant in merchant_names:
        vertical, tier, seasonality_codes = infer_vertical_and_seasonality(merchant)
        peak_months = seasonality_to_peak_months(seasonality_codes)

        engagement_months = []
        for peak in peak_months:
            engagement_months.extend(peak_month_to_pre_engagement_months(peak))

        deduped_engagement_months = []
        for m in engagement_months:
            if m not in deduped_engagement_months:
                deduped_engagement_months.append(m)

        if not deduped_engagement_months:
            deduped_engagement_months = ["2026-09"]

        for month_key in deduped_engagement_months:
            matched_peak = None
            for peak in peak_months:
                if month_key in peak_month_to_pre_engagement_months(peak):
                    matched_peak = peak
                    break
            if matched_peak is None:
                matched_peak = peak_months[0] if peak_months else "Nov"

            priority = "Critical" if month_key in ["2026-08", "2026-09", "2026-10"] and "holiday" in seasonality_codes else "High"
            if tier == "Tier 2" and priority == "High":
                priority = "Medium"

            rows.append({
                "merchant": merchant,
                "vertical": vertical,
                "tier": tier,
                "seasonality": ", ".join(seasonality_codes),
                "peak_months": ", ".join(peak_months),
                "engagement_month_label": month_to_label(month_key),
                "engagement_month_sort": month_key,
                "engagement_date": f"{month_key}-{DEFAULT_DAY}",
                "engagement_type": engagement_type_for_month(month_key, matched_peak),
                "priority": priority,
                "owner": merchant_to_owner.get(merchant, ""),
                "fy26_fc_gmv": gmv_map.get(merchant, ""),
                "status": "Planned",
                "next_action": next_action_for_priority(priority),
                "leadership_flag": leadership_flag(tier, priority),
                "playbook": playbook_for_vertical(vertical),
            })

    rows.sort(key=lambda r: (r["engagement_month_sort"], r["tier"], r["merchant"]))
    return rows

def main():
    merchant_names, merchant_to_owner, merchant_to_gmv = read_merchants_from_xlsx(INPUT_FILE)
    rows = build_rows(merchant_names, merchant_to_owner, merchant_to_gmv)

    fieldnames = [
        "merchant",
        "vertical",
        "tier",
        "seasonality",
        "peak_months",
        "engagement_month_label",
        "engagement_month_sort",
        "engagement_date",
        "engagement_type",
        "priority",
        "owner",
        "fy26_fc_gmv",
        "status",
        "next_action",
        "leadership_flag",
        "playbook",
    ]

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Loaded merchants: {len(merchant_names)}")
    print(f"Generated rows: {len(rows)}")
    print(f"CSV generated: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()